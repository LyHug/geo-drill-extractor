"""
结果导出器 - 多格式导出系统
"""

import json
import csv
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from datetime import datetime
import logging
from threading import Lock

from ..core import (
    ProcessResult,
    DrillHoleEntity, 
    Coordinate,
    ConfigLoader,
    get_config_loader,
    SingleRunMetrics,
    AggregatedMetrics,
    ExportException
)

logger = logging.getLogger(__name__)


class FieldMapper:
    """字段映射器 - 负责将数据模型映射为导出格式"""

    def __init__(self, config_loader: ConfigLoader):
        self.config_loader = config_loader

    def map_drill_hole_to_dict(self, hole: DrillHoleEntity, result: ProcessResult) -> Dict[str, Any]:
        """将钻孔实体映射为字典格式"""
        mapped_data = {
            'document_name': result.document_name,
            'processing_time': result.processing_time,
            'hole_id': hole.hole_id,
            'location_desc': hole.location_desc or '',
            'location_desc_direction_type': hole.location_desc_direction_type or '',
            'confidence': hole.confidence,
            'extracted_at': hole.extracted_at.strftime('%Y-%m-%d %H:%M:%S')
        }

        if result.metadata:
            if 'model_name' in result.metadata:
                mapped_data['model_name'] = result.metadata['model_name']
            if 'repetition_round' in result.metadata:
                mapped_data['repetition_round'] = result.metadata['repetition_round']

        # 映射坐标信息
        hole_coords = result.coordinates.get(hole.hole_id, {})
        start_coord = hole_coords.get('start')
        end_coord = hole_coords.get('end')

        if start_coord:
            mapped_data.update({
                'start_coord_x': start_coord.x,
                'start_coord_y': start_coord.y,
                'start_coord_z': start_coord.z,
                'start_coord_confidence': start_coord.confidence
            })
        else:
            mapped_data.update({
                'start_coord_x': None,
                'start_coord_y': None,
                'start_coord_z': None,
                'start_coord_confidence': None
            })

        if end_coord:
            mapped_data.update({
                'end_coord_x': end_coord.x,
                'end_coord_y': end_coord.y,
                'end_coord_z': end_coord.z,
                'end_coord_confidence': end_coord.confidence
            })
        else:
            mapped_data.update({
                'end_coord_x': None,
                'end_coord_y': None,
                'end_coord_z': None,
                'end_coord_confidence': None
            })

        # 映射实际参数
        if hole.actual_params:
            mapped_data.update({
                'actual_depth': hole.actual_params.actual_depth,
                'actual_azimuth': hole.actual_params.actual_azimuth,
                'actual_inclination': hole.actual_params.actual_inclination,
                'actual_diameter': hole.actual_params.actual_diameter,
                'start_formation': hole.actual_params.start_formation or '',
                'end_formation': hole.actual_params.end_formation or '',
                'drilling_date': hole.actual_params.drilling_date or ''
            })
        else:
            mapped_data.update({
                'actual_depth': None,
                'actual_azimuth': None,
                'actual_inclination': None,
                'actual_diameter': None,
                'start_formation': '',
                'end_formation': '',
                'drilling_date': ''
            })

        # 映射设计参数
        if hole.design_params:
            mapped_data.update({
                'design_depth': hole.design_params.design_depth,
                'design_azimuth': hole.design_params.design_azimuth,
                'design_inclination': hole.design_params.design_inclination,
                'design_diameter': hole.design_params.design_diameter,
                'design_purpose': hole.design_params.design_purpose or ''
            })
        else:
            mapped_data.update({
                'design_depth': None,
                'design_azimuth': None,
                'design_inclination': None,
                'design_diameter': None,
                'design_purpose': ''
            })

        return mapped_data
    
    def map_metrics_to_dict(self, metrics: Union[SingleRunMetrics, AggregatedMetrics]) -> Dict[str, Any]:
        """将指标对象映射为字典格式"""
        mapped_data = {
            'model_name': metrics.raw_data.model_name,
            'document_name': metrics.raw_data.document_name,
            'extracted_entities_count': metrics.raw_data.extracted_entities_count,
            'extracted_entities_with_location_count': metrics.raw_data.extracted_entities_with_location_count,
            'extracted_coordinates_count': metrics.raw_data.extracted_coordinates_count,
            'true_total_entities': metrics.raw_data.true_total_entities,
            'true_entities_with_location': metrics.raw_data.true_entities_with_location,
            'document_token_length': metrics.raw_data.document_token_length,
            'total_processing_time': metrics.raw_data.total_processing_time,
            'entity_extraction_time': metrics.raw_data.entity_extraction_time,
            'extraction_density': metrics.raw_data.extraction_density,
            'unique_location_descriptions_count': metrics.raw_data.unique_location_descriptions_count,
            'unique_location_descriptions_processing_time': metrics.raw_data.unique_location_descriptions_processing_time,
            'extraction_recall': metrics.scores.extraction_recall,
            'location_recall': metrics.scores.location_recall,
            'coordinate_success_rate': metrics.scores.coordinate_success_rate,
            'processing_stability': metrics.scores.processing_stability,
            'efficiency_coefficient': metrics.scores.efficiency_coefficient,
            'avg_location_processing_time': metrics.scores.avg_location_processing_time
        }
        
        # 聚合数据的特殊字段
        if isinstance(metrics, AggregatedMetrics):
            mapped_data.update({
                'total_repetitions': metrics.raw_data.total_repetitions,
                'aggregation_method': metrics.raw_data.aggregation_method,
                'processing_time_cv': metrics.raw_data.processing_time_cv
            })
        else:
            mapped_data['repetition_round'] = metrics.raw_data.repetition_round
            
        return mapped_data


class CSVExporter:
    """CSV格式导出器"""

    def __init__(self, config_loader: ConfigLoader):
        self.config_loader = config_loader
        self.field_mapper = FieldMapper(config_loader)

    def export_results_to_csv(self, results: List[ProcessResult], output_file: Path) -> str:
        """导出结果为CSV格式"""
        if not results:
            logger.warning("没有结果数据，创建空CSV文件")
            self._create_empty_csv(output_file)
            return str(output_file)

        csv_data = self._prepare_results_csv_data(results)
        if not csv_data:
            logger.warning("没有有效的CSV数据")
            self._create_empty_csv(output_file)
            return str(output_file)

        self._write_csv_data(csv_data, output_file)
        logger.info(f"CSV导出完成: {output_file}, 共 {len(csv_data)} 行")
        return str(output_file)
    
    def export_metrics_to_csv(self, metrics: List[Union[SingleRunMetrics, AggregatedMetrics]], output_file: Path) -> str:
        """导出指标为CSV格式"""
        if not metrics:
            logger.warning("没有指标数据，创建空CSV文件")
            self._create_empty_metrics_csv(output_file)
            return str(output_file)
            
        csv_data = []
        for metric in metrics:
            mapped_data = self.field_mapper.map_metrics_to_dict(metric)
            csv_data.append(mapped_data)
        
        if csv_data:
            df = pd.DataFrame(csv_data)
            df.to_csv(output_file, index=False, encoding='utf-8-sig')
            logger.info(f"指标CSV导出完成: {output_file}, 共 {len(csv_data)} 行")
        
        return str(output_file)

    def _prepare_results_csv_data(self, results: List[ProcessResult]) -> List[Dict[str, Any]]:
        """准备结果CSV数据"""
        csv_data = []

        for result in results:
            if not result.drill_holes:
                continue

            for hole in result.drill_holes:
                mapped_data = self.field_mapper.map_drill_hole_to_dict(hole, result)
                csv_data.append(mapped_data)

        return csv_data

    def _write_csv_data(self, csv_data: List[Dict[str, Any]], output_file: Path):
        """写入CSV数据"""
        if not csv_data:
            return
            
        # 获取所有可能的字段名
        fieldnames = set()
        for row in csv_data:
            fieldnames.update(row.keys())
        
        fieldnames = sorted(list(fieldnames))

        with open(output_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(csv_data)

    def _create_empty_csv(self, output_file: Path):
        """创建空CSV文件"""
        fieldnames = ['document_name', 'hole_id', 'location_desc', 'confidence']
        with open(output_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
    
    def _create_empty_metrics_csv(self, output_file: Path):
        """创建空指标CSV文件"""
        fieldnames = ['model_name', 'document_name', 'extraction_recall', 'coordinate_success_rate']
        with open(output_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()


class ExcelExporter:
    """Excel格式导出器"""

    def __init__(self, config_loader: ConfigLoader):
        self.config_loader = config_loader
        self.field_mapper = FieldMapper(config_loader)

    def export_results_to_excel(self, results: List[ProcessResult], output_file: Path) -> str:
        """导出结果为Excel格式"""
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.error("openpyxl 未安装，无法导出Excel格式")
            raise ImportError("请安装 openpyxl: pip install openpyxl")

        if not results:
            logger.warning("没有结果数据，创建空Excel文件")
            self._create_empty_excel(output_file)
            return str(output_file)

        wb = Workbook()
        
        # 创建主数据表
        self._create_main_results_sheet(wb, results)
        
        # 创建坐标详情表
        self._create_coord_sheet(wb, results)
        
        # 创建统计表
        self._create_stats_sheet(wb, results)

        # 删除默认工作表
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        wb.save(output_file)
        logger.info(f"Excel导出完成: {output_file}")
        return str(output_file)
    
    def export_metrics_to_excel(self, metrics: List[Union[SingleRunMetrics, AggregatedMetrics]], output_file: Path) -> str:
        """导出指标为Excel格式"""
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")
            
        if not metrics:
            self._create_empty_excel(output_file)
            return str(output_file)
            
        wb = Workbook()
        ws = wb.active
        ws.title = "指标数据"
        
        # 准备数据
        data = []
        for metric in metrics:
            mapped_data = self.field_mapper.map_metrics_to_dict(metric)
            data.append(mapped_data)
        
        if data:
            # 写入表头
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            for row, item in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=item.get(header))
        
        wb.save(output_file)
        logger.info(f"指标Excel导出完成: {output_file}")
        return str(output_file)

    def _create_main_results_sheet(self, workbook, results: List[ProcessResult]):
        """创建主数据表"""
        ws = workbook.active
        ws.title = "钻孔汇总"
        
        # 准备数据
        main_data = []
        for result in results:
            for hole in result.drill_holes:
                mapped_data = self.field_mapper.map_drill_hole_to_dict(hole, result)
                main_data.append(mapped_data)
        
        if not main_data:
            return

        # 写入表头
        headers = list(main_data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 写入数据
        for row, data in enumerate(main_data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=data.get(header))

    def _create_coord_sheet(self, workbook, results: List[ProcessResult]):
        """创建坐标详情表"""
        ws = workbook.create_sheet("坐标详情")
        
        coord_data = []
        for result in results:
            for hole_id, coords in result.coordinates.items():
                for coord_type, coord in coords.items():
                    coord_data.append({
                        'document_name': result.document_name,
                        'hole_id': hole_id,
                        'coord_type': coord_type,
                        'x_coord': coord.x,
                        'y_coord': coord.y,
                        'z_coord': coord.z,
                        'confidence': coord.confidence,
                        'method': coord.method
                    })
        
        if coord_data:
            headers = list(coord_data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                
            for row, data in enumerate(coord_data, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=data.get(header))

    def _create_stats_sheet(self, workbook, results: List[ProcessResult]):
        """创建统计表"""
        ws = workbook.create_sheet("文档统计")
        
        stats_data = []
        for result in results:
            stats_data.append({
                'document_name': result.document_name,
                'processing_time': result.processing_time,
                'drill_hole_count': len(result.drill_holes),
                'coordinate_count': len(result.coordinates),
                'error_count': len(result.errors),
                'success_rate': len(result.coordinates) / max(1, len(result.drill_holes))
            })
        
        if stats_data:
            headers = list(stats_data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                
            for row, data in enumerate(stats_data, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=data.get(header))

    def _create_empty_excel(self, output_file: Path):
        """创建空Excel文件"""
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "空数据"
            ws.cell(row=1, column=1, value="无数据")
            wb.save(output_file)
        except ImportError:
            logger.error("openpyxl 未安装")


class JSONExporter:
    """JSON格式导出器"""

    def __init__(self, config_loader: ConfigLoader):
        self.config_loader = config_loader

    def export_results_to_json(self, results: List[ProcessResult], output_file: Path) -> str:
        """导出结果为JSON格式"""
        if not results:
            logger.warning("没有结果数据，创建空JSON文件")
            self._create_empty_json(output_file)
            return str(output_file)

        json_data = self._prepare_results_json_data(results)

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(
                json_data,
                f,
                ensure_ascii=False,
                indent=2,
                default=self._json_serializer
            )

        logger.info(f"JSON导出完成: {output_file}")
        return str(output_file)
    
    def export_metrics_to_json(self, metrics: List[Union[SingleRunMetrics, AggregatedMetrics]], output_file: Path) -> str:
        """导出指标为JSON格式"""
        if not metrics:
            self._create_empty_metrics_json(output_file)
            return str(output_file)
            
        json_data = {
            'export_info': {
                'export_time': datetime.now().isoformat(),
                'total_metrics': len(metrics),
                'single_run_count': sum(1 for m in metrics if isinstance(m, SingleRunMetrics)),
                'aggregated_count': sum(1 for m in metrics if isinstance(m, AggregatedMetrics))
            },
            'metrics': []
        }
        
        for metric in metrics:
            mapped_data = self.field_mapper.map_metrics_to_dict(metric)
            json_data['metrics'].append(mapped_data)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2, default=self._json_serializer)
            
        logger.info(f"指标JSON导出完成: {output_file}")
        return str(output_file)

    def _prepare_results_json_data(self, results: List[ProcessResult]) -> Dict[str, Any]:
        """准备结果JSON数据"""
        documents = []
        for result in results:
            doc_data = {
                'document_name': result.document_name,
                'processing_time': result.processing_time,
                'drill_holes': [],
                'coordinates': {},
                'errors': result.errors,
                'metadata': result.metadata
            }
            
            # 添加钻孔数据
            for hole in result.drill_holes:
                hole_data = self._build_hole_data(hole)
                doc_data['drill_holes'].append(hole_data)
            
            # 添加坐标数据
            for hole_id, coords in result.coordinates.items():
                doc_data['coordinates'][hole_id] = {}
                for coord_type, coord in coords.items():
                    doc_data['coordinates'][hole_id][coord_type] = self._build_coordinate_data(coord)
            
            documents.append(doc_data)

        json_data = {
            'export_info': {
                'export_time': datetime.now().isoformat(),
                'total_documents': len(results),
                'total_drill_holes': sum(len(r.drill_holes) for r in results),
                'total_coordinates': sum(len(r.coordinates) for r in results)
            },
            'documents': documents
        }

        return json_data

    def _build_hole_data(self, hole: DrillHoleEntity) -> Dict[str, Any]:
        """构建钻孔数据"""
        hole_data = {
            'hole_id': hole.hole_id,
            'location_desc': hole.location_desc,
            'location_desc_direction_type': hole.location_desc_direction_type,
            'confidence': hole.confidence,
            'extracted_at': hole.extracted_at.isoformat() if hole.extracted_at else None
        }

        # 设计参数
        if hole.design_params:
            hole_data['design_params'] = {
                'design_depth': hole.design_params.design_depth,
                'design_azimuth': hole.design_params.design_azimuth,
                'design_inclination': hole.design_params.design_inclination,
                'design_diameter': hole.design_params.design_diameter,
                'design_purpose': hole.design_params.design_purpose
            }

        # 实际参数
        if hole.actual_params:
            hole_data['actual_params'] = {
                'actual_depth': hole.actual_params.actual_depth,
                'actual_azimuth': hole.actual_params.actual_azimuth,
                'actual_inclination': hole.actual_params.actual_inclination,
                'actual_diameter': hole.actual_params.actual_diameter,
                'start_formation': hole.actual_params.start_formation,
                'end_formation': hole.actual_params.end_formation,
                'drilling_date': hole.actual_params.drilling_date
            }

        return hole_data

    def _build_coordinate_data(self, coord: Coordinate) -> Dict[str, Any]:
        """构建坐标数据"""
        return {
            'x': coord.x,
            'y': coord.y,
            'z': coord.z,
            'confidence': coord.confidence,
            'method': coord.method
        }

    def _json_serializer(self, obj):
        """JSON序列化器，处理特殊类型"""
        if isinstance(obj, datetime):
            return obj.isoformat()
        elif hasattr(obj, 'to_dict'):
            return obj.to_dict()
        elif hasattr(obj, '__dict__'):
            return obj.__dict__
        else:
            return str(obj)

    def _create_empty_json(self, output_file: Path):
        """创建空JSON文件"""
        empty_data = {
            'export_info': {
                'export_time': datetime.now().isoformat(),
                'total_documents': 0,
                'total_drill_holes': 0,
                'total_coordinates': 0
            },
            'documents': []
        }

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(empty_data, f, ensure_ascii=False, indent=2)
    
    def _create_empty_metrics_json(self, output_file: Path):
        """创建空指标JSON文件"""
        empty_data = {
            'export_info': {
                'export_time': datetime.now().isoformat(),
                'total_metrics': 0,
                'single_run_count': 0,
                'aggregated_count': 0
            },
            'metrics': []
        }
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(empty_data, f, ensure_ascii=False, indent=2)


class ResultExporter:
    """统一结果导出器"""

    def __init__(self, config_loader: Optional[ConfigLoader] = None):
        self.config_loader = config_loader or get_config_loader()
        
        # 初始化各格式导出器
        self.csv_exporter = CSVExporter(self.config_loader)
        self.excel_exporter = ExcelExporter(self.config_loader)
        self.json_exporter = JSONExporter(self.config_loader)
        
        # 线程安全锁
        self._export_lock = Lock()

    def export_results(
        self,
        results: List[ProcessResult],
        output_dir: Path,
        formats: List[str] = None,
        filename_prefix: str = "drill_holes"
    ) -> Dict[str, str]:
        """
        导出处理结果
        
        Args:
            results: 处理结果列表
            output_dir: 输出目录
            formats: 导出格式列表，默认为['csv', 'excel', 'json']
            filename_prefix: 文件名前缀
            
        Returns:
            格式到文件路径的映射
        """
        if formats is None:
            formats = ['csv', 'excel', 'json']
        
        with self._export_lock:
            output_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            exported_files = {}
            
            for format_type in formats:
                try:
                    if format_type.lower() == 'csv':
                        output_file = output_dir / f"{filename_prefix}_{timestamp}.csv"
                        exported_files['csv'] = self.csv_exporter.export_results_to_csv(results, output_file)
                    elif format_type.lower() == 'excel':
                        output_file = output_dir / f"{filename_prefix}_{timestamp}.xlsx"
                        exported_files['excel'] = self.excel_exporter.export_results_to_excel(results, output_file)
                    elif format_type.lower() == 'json':
                        output_file = output_dir / f"{filename_prefix}_{timestamp}.json"
                        exported_files['json'] = self.json_exporter.export_results_to_json(results, output_file)
                    else:
                        logger.warning(f"不支持的导出格式: {format_type}")
                        continue
                        
                    logger.info(f"成功导出 {format_type.upper()} 格式: {exported_files[format_type.lower()]}")
                    
                except Exception as e:
                    logger.error(f"导出 {format_type} 格式失败: {str(e)}")
                    continue
            
            return exported_files
    
    def export_metrics(
        self,
        metrics: List[Union[SingleRunMetrics, AggregatedMetrics]],
        output_dir: Path,
        formats: List[str] = None,
        filename_prefix: str = "metrics"
    ) -> Dict[str, str]:
        """
        导出评估指标
        
        Args:
            metrics: 指标列表
            output_dir: 输出目录
            formats: 导出格式列表
            filename_prefix: 文件名前缀
            
        Returns:
            格式到文件路径的映射
        """
        if formats is None:
            formats = ['csv', 'excel', 'json']
            
        with self._export_lock:
            output_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            exported_files = {}
            
            for format_type in formats:
                try:
                    if format_type.lower() == 'csv':
                        output_file = output_dir / f"{filename_prefix}_{timestamp}.csv"
                        exported_files['csv'] = self.csv_exporter.export_metrics_to_csv(metrics, output_file)
                    elif format_type.lower() == 'excel':
                        output_file = output_dir / f"{filename_prefix}_{timestamp}.xlsx"
                        exported_files['excel'] = self.excel_exporter.export_metrics_to_excel(metrics, output_file)
                    elif format_type.lower() == 'json':
                        output_file = output_dir / f"{filename_prefix}_{timestamp}.json"
                        exported_files['json'] = self.json_exporter.export_metrics_to_json(metrics, output_file)
                    else:
                        logger.warning(f"不支持的导出格式: {format_type}")
                        continue
                        
                    logger.info(f"成功导出指标 {format_type.upper()} 格式: {exported_files[format_type.lower()]}")
                    
                except Exception as e:
                    logger.error(f"导出指标 {format_type} 格式失败: {str(e)}")
                    continue
            
            return exported_files

    def export_single_result(
        self,
        result: ProcessResult,
        output_dir: Path,
        format_type: str
    ) -> str:
        """按格式导出单个结果"""
        with self._export_lock:
            output_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            doc_name = result.document_name.replace('.docx', '').replace('.', '_')

            if format_type.lower() == 'csv':
                output_file = output_dir / f"{doc_name}_{timestamp}.csv"
                return self.csv_exporter.export_results_to_csv([result], output_file)
            elif format_type.lower() == 'excel':
                output_file = output_dir / f"{doc_name}_{timestamp}.xlsx"
                return self.excel_exporter.export_results_to_excel([result], output_file)
            elif format_type.lower() == 'json':
                output_file = output_dir / f"{doc_name}_{timestamp}.json"
                return self.json_exporter.export_results_to_json([result], output_file)
            else:
                raise ValueError(f"不支持的导出格式: {format_type}")

    def get_export_summary(self) -> Dict[str, Any]:
        """获取导出器摘要"""
        return {
            'supported_formats': ['csv', 'excel', 'json'],
            'config_loaded': self.config_loader is not None,
            'exporters': {
                'csv': 'CSVExporter',
                'excel': 'ExcelExporter', 
                'json': 'JSONExporter'
            }
        }